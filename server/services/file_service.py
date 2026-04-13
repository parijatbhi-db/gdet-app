"""File generation for CSV, JSON, and XLSX formats."""
import csv
import io
import json
from datetime import datetime
from typing import Any


def generate_file(
    columns: list[str],
    rows: list[list[Any]],
    config: dict,
    definition: dict,
) -> tuple[bytes, str]:
    """Generate an output file. Returns (file_bytes, filename)."""
    fmt = config.get("file_format", "csv").lower()
    filename = _build_filename(config, definition, fmt)

    if fmt == "csv":
        file_bytes = _generate_csv(columns, rows, config)
    elif fmt == "json":
        file_bytes = _generate_json(columns, rows, config)
    elif fmt == "xlsx":
        file_bytes = _generate_xlsx(columns, rows, config)
    else:
        file_bytes = _generate_csv(columns, rows, config)

    return file_bytes, filename


def _generate_csv(columns: list[str], rows: list[list[Any]], config: dict) -> bytes:
    delimiter = config.get("delimiter", ",")
    encoding = config.get("encoding", "utf-8")
    decimal_format = config.get("decimal_format", ".")

    output = io.StringIO()
    writer = csv.writer(output, delimiter=delimiter)
    writer.writerow(columns)
    for row in rows:
        processed = []
        for val in row:
            if val is None:
                processed.append("")
            elif decimal_format == "," and _is_numeric(val):
                processed.append(str(val).replace(".", ","))
            else:
                processed.append(str(val))
        writer.writerow(processed)

    return output.getvalue().encode(encoding)


def _generate_json(columns: list[str], rows: list[list[Any]], config: dict) -> bytes:
    encoding = config.get("encoding", "utf-8")
    data = [dict(zip(columns, row)) for row in rows]
    return json.dumps(data, indent=2, ensure_ascii=False, default=str).encode(encoding)


def _generate_xlsx(columns: list[str], rows: list[list[Any]], config: dict) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Extract Data"

    # Header style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="005DA6", end_color="005DA6", fill_type="solid")

    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill

    for row_idx, row in enumerate(rows, 2):
        for col_idx, val in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    # Auto-width columns
    for col_idx, col_name in enumerate(columns, 1):
        max_len = len(str(col_name))
        for row in rows[:100]:
            if col_idx - 1 < len(row) and row[col_idx - 1]:
                max_len = max(max_len, len(str(row[col_idx - 1])))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 2, 50)

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def _build_filename(config: dict, definition: dict, fmt: str) -> str:
    template = config.get("file_naming_template", "{name}_{date}")
    date_str = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    name_clean = definition.get("name", "extract").replace(" ", "_").lower()

    filename = template.replace("{date}", date_str)
    filename = filename.replace("{name}", name_clean)
    filename = filename.replace("{country}", definition.get("country_code", "") or "")
    filename = filename.replace("{type}", definition.get("extract_type", "") or "")

    ext_map = {"csv": ".csv", "json": ".json", "xlsx": ".xlsx"}
    return f"{filename}{ext_map.get(fmt, '.csv')}"


def _is_numeric(val) -> bool:
    try:
        float(val)
        return True
    except (ValueError, TypeError):
        return False
